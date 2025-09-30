"""
Comando Django para mapear usu√°rios √∫nicos de todas as planilhas extra√≠das.
Elimina duplica√ß√µes e cria relat√≥rio consolidado com emails e pap√©is corretos.
"""

import json
import re
import openpyxl
from pathlib import Path
from collections import defaultdict, Counter
from django.core.management.base import BaseCommand
from django.contrib.auth.models import Group
from django.db import transaction
from core.models import Usuario, Setor


class Command(BaseCommand):
    help = 'Mapeia usu√°rios √∫nicos de todas as planilhas, eliminando duplica√ß√µes'

    def add_arguments(self, parser):
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Executa em modo simula√ß√£o sem fazer altera√ß√µes'
        )
        parser.add_argument(
            '--verbose',
            action='store_true',
            help='Mostra informa√ß√µes detalhadas do processo'
        )
        parser.add_argument(
            '--output-file',
            type=str,
            default='relatorio_usuarios_unicos.json',
            help='Nome do arquivo de sa√≠da para o relat√≥rio'
        )
        parser.add_argument(
            '--planilhas-dir',
            type=str,
            default='archive/spreadsheets',
            help='Diret√≥rio das planilhas Excel'
        )

    def handle(self, *args, **options):
        self.dry_run = options['dry_run']
        self.verbose = options['verbose']
        self.output_file = options['output_file']
        self.planilhas_dir = Path(options['planilhas_dir'])

        if self.dry_run:
            self.stdout.write(
                self.style.WARNING('MODO SIMULA√á√ÉO - Nenhuma altera√ß√£o ser√° feita')
            )

        # Estrutura para armazenar usu√°rios √∫nicos
        self.usuarios_unicos = {}
        self.duplicatas_encontradas = []
        self.problemas_dados = []
        self.estatisticas = {
            'total_registros_processados': 0,
            'usuarios_unicos_encontrados': 0,
            'duplicatas_eliminadas': 0,
            'emails_faltantes': 0,
            'cpfs_invalidos': 0,
            'nomes_problematicos': 0
        }

        # Mapeamento de cargos para grupos Django
        self.cargo_para_grupo = {
            'Coordenadores': 'coordenador',
            'Superintend√™ncia': 'superintendencia',
            'Gerentes': 'superintendencia',
            'Controle': 'controle',
            'Formadores': 'formador',
            'Diretoria': 'diretoria',
            'Admin': 'admin'
        }

        # Mapeamento de projetos para setores
        self.projeto_para_setor = {
            'ACerta': 'ACerta',
            'Brincando e Aprendendo': 'Brincando e Aprendendo',
            'Vidas': 'Vida & Linguagem',
            'Vida & Linguagem': 'Vida & Linguagem',
            'Vida & Matem√°tica': 'Vida & Matem√°tica',
            'IDEB10': 'IDEB10',
            'Fluir das Emo√ß√µes': 'Fluir das Emo√ß√µes',
            'AMMA': 'AMMA',
            'Ler, Ouvir e Contar': 'Ler, Ouvir e Contar',
            'Superintend√™ncia': 'Superintend√™ncia'
        }

        self.stdout.write(f"\nüîç INICIANDO MAPEAMENTO DE USU√ÅRIOS √öNICOS")
        self.stdout.write(f"   Diret√≥rio: {self.planilhas_dir}")
        self.stdout.write(f"   Arquivo de sa√≠da: {self.output_file}")

        # Processar todas as planilhas
        self._processar_planilhas()

        # Analisar dados da agenda principal (formadores identificados)
        self._processar_dados_agenda()

        # Eliminar duplicatas
        self._eliminar_duplicatas()

        # Gerar relat√≥rio
        self._gerar_relatorio()

        # Criar/atualizar usu√°rios no sistema (se n√£o for dry-run)
        if not self.dry_run:
            self._criar_usuarios_sistema()

        # Estat√≠sticas finais
        self._exibir_estatisticas()

    def _processar_planilhas(self):
        """Processa todas as planilhas Excel no diret√≥rio"""
        planilhas_encontradas = list(self.planilhas_dir.glob("*.xlsx"))

        if not planilhas_encontradas:
            self.stdout.write(
                self.style.ERROR(f'Nenhuma planilha encontrada em {self.planilhas_dir}')
            )
            return

        self.stdout.write(f"\nüìä PROCESSANDO {len(planilhas_encontradas)} PLANILHAS:")

        for planilha_path in planilhas_encontradas:
            if planilha_path.name == 'produtos.xlsx':
                continue  # Pular planilha de produtos

            self.stdout.write(f"   üìÑ {planilha_path.name}")
            try:
                self._processar_planilha_excel(planilha_path)
            except Exception as e:
                self.stdout.write(
                    self.style.ERROR(f"      Erro ao processar {planilha_path.name}: {str(e)}")
                )

    def _processar_planilha_excel(self, planilha_path):
        """Processa uma planilha Excel espec√≠fica"""
        workbook = openpyxl.load_workbook(planilha_path, read_only=True)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Ler header (primeira linha)
            headers = []
            for cell in sheet[1]:
                headers.append(cell.value or '')

            # Verificar se tem estrutura de usu√°rios
            if not self._tem_estrutura_usuarios(headers):
                continue

            # Mapear √≠ndices das colunas importantes
            indices = self._mapear_indices_colunas(headers)
            if not indices:
                continue

            projeto_nome = self._extrair_nome_projeto(planilha_path.name)

            # Processar linhas de dados
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
                if not row or len(row) <= max(indices.values()):
                    continue

                usuario_data = self._extrair_dados_usuario(row, indices, projeto_nome)
                if usuario_data:
                    self._adicionar_usuario_unico(usuario_data)
                    self.estatisticas['total_registros_processados'] += 1

        workbook.close()

    def _tem_estrutura_usuarios(self, headers):
        """Verifica se a planilha tem estrutura de usu√°rios"""
        headers_lower = [h.lower() for h in headers if h]
        required_fields = ['nome', 'email', 'cargo']
        return any(field in ' '.join(headers_lower) for field in required_fields)

    def _mapear_indices_colunas(self, headers):
        """Mapeia √≠ndices das colunas importantes"""
        indices = {}
        headers_lower = [h.lower() if h else '' for h in headers]

        for i, header in enumerate(headers_lower):
            if 'nome' in header and 'completo' not in header:
                indices['nome'] = i
            elif 'nome completo' in header:
                indices['nome_completo'] = i
            elif 'cpf' in header:
                indices['cpf'] = i
            elif 'email' in header:
                indices['email'] = i
            elif 'cargo' in header:
                indices['cargo'] = i
            elif 'grupo' in header:
                indices['grupo'] = i
            elif 'projeto' in header:
                indices['projeto'] = i

        return indices if 'nome' in indices else None

    def _extrair_nome_projeto(self, nome_arquivo):
        """Extrai nome do projeto baseado no nome do arquivo"""
        nome_base = nome_arquivo.replace('.xlsx', '')
        return self.projeto_para_setor.get(nome_base, nome_base)

    def _extrair_dados_usuario(self, row, indices, projeto_nome):
        """Extrai dados de usu√°rio de uma linha da planilha"""
        try:
            # Extrair campos b√°sicos
            nome = row[indices['nome']] if indices.get('nome') is not None else ''
            nome_completo = row[indices.get('nome_completo', indices['nome'])] if indices.get('nome_completo') is not None else nome
            cpf = row[indices['cpf']] if indices.get('cpf') is not None else ''
            email = row[indices['email']] if indices.get('email') is not None else ''
            cargo = row[indices['cargo']] if indices.get('cargo') is not None else ''
            grupo = row[indices.get('grupo', indices['cargo'])] if indices.get('grupo') is not None else cargo
            projeto = row[indices['projeto']] if indices.get('projeto') is not None else projeto_nome

            # Normalizar dados
            nome = str(nome).strip() if nome else ''
            nome_completo = str(nome_completo).strip() if nome_completo else nome
            cpf = self._normalizar_cpf(str(cpf) if cpf else '')
            email = str(email).strip().lower() if email else ''
            cargo = str(cargo).strip() if cargo else ''
            grupo = str(grupo).strip() if grupo else cargo
            projeto = str(projeto).strip() if projeto else projeto_nome

            # Valida√ß√µes b√°sicas
            if not nome or len(nome) < 2:
                return None

            # Verificar problemas nos dados
            problemas = []
            if '?' in nome:
                problemas.append(f"Nome question√°vel: {nome}")
                self.estatisticas['nomes_problematicos'] += 1

            if nome.upper() == 'SOLICITADO':
                problemas.append("Nome 'SOLICITADO' - formador n√£o definido")
                self.estatisticas['nomes_problematicos'] += 1

            if not email:
                problemas.append("Email faltante")
                self.estatisticas['emails_faltantes'] += 1

            if cpf and len(cpf) != 11:
                problemas.append(f"CPF inv√°lido: {cpf}")
                self.estatisticas['cpfs_invalidos'] += 1

            # Mapear cargo para grupo Django
            grupo_django = self.cargo_para_grupo.get(cargo, self.cargo_para_grupo.get(grupo, 'formador'))

            usuario_data = {
                'nome': nome,
                'nome_completo': nome_completo,
                'cpf': cpf,
                'email': email,
                'cargo_original': cargo,
                'grupo_original': grupo,
                'grupo_django': grupo_django,
                'projeto': projeto,
                'setor': self.projeto_para_setor.get(projeto, projeto),
                'fonte_planilha': projeto_nome,
                'problemas': problemas
            }

            if problemas and self.verbose:
                for problema in problemas:
                    self.stdout.write(f"      ‚ö†Ô∏è  {problema}")

            return usuario_data

        except Exception as e:
            if self.verbose:
                self.stdout.write(f"      ‚ùå Erro ao processar linha: {str(e)}")
            return None

    def _normalizar_cpf(self, cpf):
        """Normaliza CPF removendo formata√ß√£o"""
        if not cpf:
            return ''
        cpf_numeros = re.sub(r'[^\d]', '', str(cpf))
        return cpf_numeros if len(cpf_numeros) == 11 else ''

    def _processar_dados_agenda(self):
        """Processa dados da agenda principal para adicionar formadores identificados"""
        try:
            # Ler dados da an√°lise da agenda principal
            agenda_file = Path('ANALISE_COMPLETA_PLANILHA_AGENDA_2025.md')
            if agenda_file.exists():
                with open(agenda_file, 'r', encoding='utf-8') as f:
                    conteudo = f.read()

                # Extrair lista de formadores da an√°lise
                formadores_agenda = self._extrair_formadores_da_agenda(conteudo)

                for formador_nome in formadores_agenda:
                    # Verificar se j√° n√£o est√° mapeado
                    if not self._usuario_ja_existe(formador_nome):
                        usuario_data = {
                            'nome': formador_nome,
                            'nome_completo': formador_nome,
                            'cpf': '',
                            'email': '',
                            'cargo_original': 'Formador',
                            'grupo_original': 'Formadores',
                            'grupo_django': 'formador',
                            'projeto': 'M√∫ltiplos Projetos',
                            'setor': 'Forma√ß√£o',
                            'fonte_planilha': 'Agenda Principal',
                            'problemas': ['Dados incompletos - apenas nome identificado']
                        }
                        self._adicionar_usuario_unico(usuario_data)

        except Exception as e:
            if self.verbose:
                self.stdout.write(f"   ‚ö†Ô∏è  Erro ao processar dados da agenda: {str(e)}")

    def _extrair_formadores_da_agenda(self, conteudo):
        """Extrai lista de formadores da an√°lise da agenda"""
        formadores = []
        # Procurar se√ß√£o de formadores
        linhas = conteudo.split('\n')
        em_secao_formadores = False

        for linha in linhas:
            if 'FORMADORES IDENTIFICADOS' in linha:
                em_secao_formadores = True
                continue
            elif em_secao_formadores and linha.startswith('### '):
                break
            elif em_secao_formadores and linha.strip().startswith('- '):
                # Extrair nome do formador
                nome = linha.strip()[2:].strip()
                if nome and len(nome) > 2:
                    formadores.append(nome)

        return formadores

    def _usuario_ja_existe(self, nome):
        """Verifica se usu√°rio j√° existe na lista de √∫nicos"""
        for usuario_id, usuario_data in self.usuarios_unicos.items():
            if usuario_data['nome'].lower() == nome.lower():
                return True
        return False

    def _adicionar_usuario_unico(self, usuario_data):
        """Adiciona usu√°rio √† lista de √∫nicos, verificando duplicatas"""
        # Gerar chave √∫nica baseada em CPF > Email > Nome
        chave = self._gerar_chave_usuario(usuario_data)

        if chave in self.usuarios_unicos:
            # Detectou duplicata - mesclar dados
            self._mesclar_usuario_duplicato(chave, usuario_data)
            self.estatisticas['duplicatas_eliminadas'] += 1
        else:
            # Novo usu√°rio √∫nico
            self.usuarios_unicos[chave] = usuario_data
            self.estatisticas['usuarios_unicos_encontrados'] += 1

    def _gerar_chave_usuario(self, usuario_data):
        """Gera chave √∫nica para identifica√ß√£o de usu√°rio"""
        # Prioridade: CPF > Email > Nome normalizado
        if usuario_data['cpf']:
            return f"cpf_{usuario_data['cpf']}"
        elif usuario_data['email']:
            return f"email_{usuario_data['email']}"
        else:
            nome_normalizado = re.sub(r'[^\w\s]', '', usuario_data['nome']).lower().replace(' ', '_')
            return f"nome_{nome_normalizado}"

    def _mesclar_usuario_duplicato(self, chave, novo_usuario):
        """Mescla dados de usu√°rio duplicado"""
        usuario_existente = self.usuarios_unicos[chave]

        # Mesclar informa√ß√µes (priorizar dados mais completos)
        if not usuario_existente['email'] and novo_usuario['email']:
            usuario_existente['email'] = novo_usuario['email']

        if not usuario_existente['cpf'] and novo_usuario['cpf']:
            usuario_existente['cpf'] = novo_usuario['cpf']

        if len(novo_usuario['nome_completo']) > len(usuario_existente['nome_completo']):
            usuario_existente['nome_completo'] = novo_usuario['nome_completo']

        # Adicionar projetos m√∫ltiplos
        if novo_usuario['projeto'] not in usuario_existente['projeto']:
            usuario_existente['projeto'] += f", {novo_usuario['projeto']}"

        # Combinar problemas
        usuario_existente['problemas'].extend(novo_usuario['problemas'])

        # Registrar duplicata encontrada
        self.duplicatas_encontradas.append({
            'chave': chave,
            'nome': usuario_existente['nome'],
            'fontes': [usuario_existente['fonte_planilha'], novo_usuario['fonte_planilha']]
        })

    def _eliminar_duplicatas(self):
        """Fase final de elimina√ß√£o de duplicatas por similaridade de nomes"""
        self.stdout.write(f"\nüîÑ ELIMINANDO DUPLICATAS POR SIMILARIDADE...")

        usuarios_list = list(self.usuarios_unicos.items())
        duplicatas_por_nome = []

        for i, (chave1, usuario1) in enumerate(usuarios_list):
            for j, (chave2, usuario2) in enumerate(usuarios_list[i+1:], i+1):
                if self._nomes_similares(usuario1['nome'], usuario2['nome']):
                    duplicatas_por_nome.append((chave1, chave2))

        # Processar duplicatas encontradas
        for chave1, chave2 in duplicatas_por_nome:
            if chave1 in self.usuarios_unicos and chave2 in self.usuarios_unicos:
                self.stdout.write(f"   üîó Mesclando: {self.usuarios_unicos[chave1]['nome']} ‚Üî {self.usuarios_unicos[chave2]['nome']}")
                self._mesclar_usuario_duplicato(chave1, self.usuarios_unicos[chave2])
                del self.usuarios_unicos[chave2]
                self.estatisticas['duplicatas_eliminadas'] += 1

    def _nomes_similares(self, nome1, nome2):
        """Verifica se dois nomes s√£o similares (mesmo primeiro e √∫ltimo nome)"""
        def normalizar_nome(nome):
            return re.sub(r'[^\w\s]', '', nome).lower().split()

        palavras1 = normalizar_nome(nome1)
        palavras2 = normalizar_nome(nome2)

        if len(palavras1) < 2 or len(palavras2) < 2:
            return False

        # Verificar se primeiro e √∫ltimo nome s√£o iguais
        return (palavras1[0] == palavras2[0] and palavras1[-1] == palavras2[-1])

    def _gerar_relatorio(self):
        """Gera relat√≥rio consolidado em JSON"""
        self.stdout.write(f"\nüìã GERANDO RELAT√ìRIO CONSOLIDADO...")

        from datetime import datetime
        relatorio = {
            'timestamp': datetime.now().isoformat(),
            'estatisticas': self.estatisticas,
            'distribuicao_por_projeto': self._calcular_distribuicao_projeto(),
            'distribuicao_por_cargo': self._calcular_distribuicao_cargo(),
            'usuarios_unicos': list(self.usuarios_unicos.values()),
            'duplicatas_encontradas': self.duplicatas_encontradas,
            'problemas_dados': self._consolidar_problemas()
        }

        with open(self.output_file, 'w', encoding='utf-8') as f:
            json.dump(relatorio, f, ensure_ascii=False, indent=2)

        self.stdout.write(f"   ‚úÖ Relat√≥rio salvo em: {self.output_file}")

    def _calcular_distribuicao_projeto(self):
        """Calcula distribui√ß√£o de usu√°rios por projeto"""
        distribuicao = Counter()
        for usuario in self.usuarios_unicos.values():
            projetos = usuario['projeto'].split(', ')
            for projeto in projetos:
                distribuicao[projeto.strip()] += 1
        return dict(distribuicao)

    def _calcular_distribuicao_cargo(self):
        """Calcula distribui√ß√£o de usu√°rios por cargo"""
        distribuicao = Counter()
        for usuario in self.usuarios_unicos.values():
            distribuicao[usuario['grupo_django']] += 1
        return dict(distribuicao)

    def _consolidar_problemas(self):
        """Consolida todos os problemas encontrados"""
        problemas = defaultdict(list)
        for usuario in self.usuarios_unicos.values():
            for problema in usuario['problemas']:
                problemas[problema].append(usuario['nome'])
        return dict(problemas)

    def _criar_usuarios_sistema(self):
        """Cria/atualiza usu√°rios no sistema Django"""
        self.stdout.write(f"\nüë• CRIANDO/ATUALIZANDO USU√ÅRIOS NO SISTEMA...")

        # Garantir que grupos existem
        grupos_necessarios = ['coordenador', 'superintendencia', 'controle', 'formador', 'diretoria', 'admin']
        for grupo_nome in grupos_necessarios:
            Group.objects.get_or_create(name=grupo_nome)

        criados = 0
        atualizados = 0
        erros = 0

        with transaction.atomic():
            for usuario_data in self.usuarios_unicos.values():
                try:
                    usuario, created = self._criar_ou_atualizar_usuario(usuario_data)
                    if created:
                        criados += 1
                    else:
                        atualizados += 1

                except Exception as e:
                    erros += 1
                    if self.verbose:
                        self.stdout.write(f"   ‚ùå Erro ao processar {usuario_data['nome']}: {str(e)}")

        self.stdout.write(f"   ‚úÖ Criados: {criados}, Atualizados: {atualizados}, Erros: {erros}")

    def _criar_ou_atualizar_usuario(self, usuario_data):
        """Cria ou atualiza um usu√°rio no sistema"""
        # Buscar usu√°rio existente
        usuario = None

        # 1. Buscar por CPF
        if usuario_data['cpf']:
            try:
                usuario = Usuario.objects.get(cpf=usuario_data['cpf'])
            except Usuario.DoesNotExist:
                pass

        # 2. Buscar por email
        if not usuario and usuario_data['email']:
            try:
                usuario = Usuario.objects.get(email=usuario_data['email'])
            except Usuario.DoesNotExist:
                pass

        # 3. Criar novo usu√°rio
        if not usuario:
            usuario = Usuario()
            created = True
        else:
            created = False

        # Atualizar dados
        if not usuario.email:
            usuario.email = usuario_data['email']

        if not usuario.cpf:
            usuario.cpf = usuario_data['cpf']

        # Separar nome
        nome_parts = usuario_data['nome_completo'].split() if usuario_data['nome_completo'] else usuario_data['nome'].split()
        if nome_parts:
            usuario.first_name = nome_parts[0]
            if len(nome_parts) > 1:
                usuario.last_name = ' '.join(nome_parts[1:])

        # Username
        if not usuario.username:
            if usuario_data['email']:
                username_base = usuario_data['email'].split('@')[0]
            else:
                username_base = usuario_data['nome'].lower().replace(' ', '')

            counter = 1
            username = username_base
            while Usuario.objects.filter(username=username).exists():
                username = f"{username_base}{counter}"
                counter += 1
            usuario.username = username

        # Cargo e grupo
        usuario.cargo = self._mapear_cargo_sistema(usuario_data['grupo_django'])

        if usuario_data['grupo_django'] == 'formador':
            usuario.formador_ativo = True

        usuario.save()

        # Adicionar ao grupo
        grupo = Group.objects.get(name=usuario_data['grupo_django'])
        usuario.groups.add(grupo)

        return usuario, created

    def _mapear_cargo_sistema(self, grupo_django):
        """Mapeia grupo Django para campo cargo"""
        mapeamento = {
            'coordenador': 'coordenador',
            'superintendencia': 'gerente',
            'controle': 'controle',
            'formador': 'formador',
            'diretoria': 'outros',
            'admin': 'admin'
        }
        return mapeamento.get(grupo_django, 'outros')

    def _exibir_estatisticas(self):
        """Exibe estat√≠sticas finais"""
        self.stdout.write(f"\nüìä ESTAT√çSTICAS FINAIS:")
        self.stdout.write(f"   Total de registros processados: {self.estatisticas['total_registros_processados']}")
        self.stdout.write(f"   Usu√°rios √∫nicos encontrados: {self.estatisticas['usuarios_unicos_encontrados']}")
        self.stdout.write(f"   Duplicatas eliminadas: {self.estatisticas['duplicatas_eliminadas']}")
        self.stdout.write(f"   Emails faltantes: {self.estatisticas['emails_faltantes']}")
        self.stdout.write(f"   CPFs inv√°lidos: {self.estatisticas['cpfs_invalidos']}")
        self.stdout.write(f"   Nomes problem√°ticos: {self.estatisticas['nomes_problematicos']}")

        self.stdout.write(f"\nüìã DISTRIBUI√á√ÉO POR CARGO:")
        distribuicao_cargo = self._calcular_distribuicao_cargo()
        for cargo, quantidade in sorted(distribuicao_cargo.items()):
            self.stdout.write(f"   {cargo}: {quantidade}")

        self.stdout.write(f"\nüéØ RELAT√ìRIO SALVO EM: {self.output_file}")

        if self.dry_run:
            self.stdout.write(
                self.style.WARNING('\n‚ö†Ô∏è  MODO SIMULA√á√ÉO - Execute sem --dry-run para aplicar as altera√ß√µes')
            )
        else:
            self.stdout.write(
                self.style.SUCCESS('\n‚úÖ MAPEAMENTO CONCLU√çDO COM SUCESSO!')
            )